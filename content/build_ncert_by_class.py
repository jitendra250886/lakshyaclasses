"""
###############       Prompt (Full & Final) : 
You are a senior Python engineer. Write a production‑ready Python 3.11 script named `build_ncert_by_class_full.py` that creates ONE Excel workbook for NCERT chapters using only the `openpyxl` library (no other external deps).
========================
GOAL & LAYOUT (MANDATORY)
========================
• The workbook must contain exactly 12 worksheets, one per class, named: “Class 01”, “Class 02”, …, “Class 12”.
• Each worksheet must contain a SINGLE Excel Table with the columns, in this exact order:
  1) Subject
  2) Chapter Index
  3) Status
  4) Remark
• Each row represents one subject of that class.
• Column “Chapter Index” must contain a **newline‑separated** list of chapter slugs in the format:
  CHAPTER-XX-TITLE-WITH-HYPHENS
  Rules:
  - If the source chapter entry is like `LESSON-XX: Title`, use that numeric `XX`.
  - Otherwise, number by the order (01, 02, …).
  - English titles: keep A–Z, 0–9; spaces → hyphens; uppercase; strip punctuation.
  - Hindi titles: **keep Devanagari characters**, convert spaces/punctuation to hyphens; do NOT transliterate; no uppercase conversion.
• Column “Status” must be a data‑validation dropdown with exactly three values:
  Not_start, working, complete
• Column “Remark” is free text.

========================
DATA (EMBED IN SCRIPT)
========================
• Embed a large in‑script constant `CURRICULUM: Dict[str, Dict[str, List[str]]]`.
  Key = "Class X" (X = 1..12).
  Value = dict of subject → list[str], where each string is either
    "LESSON-XX: Title" or just "Title".
• Populate REAL, NCERT‑accurate chapter lists for **core subjects**:
  - Classes 1–5: English, Hindi, Maths, EVS (as per commonly used NCERT editions).
  - Classes 6–8: English, Hindi, Maths, Science, SocialScience_History, SocialScience_Geography, SocialScience_Civics.
  - Classes 9–10: English, Hindi, Maths, Science, SocialScience_History, SocialScience_Geography, SocialScience_Civics, SocialScience_Economics.
  - Classes 11–12 (Science stream): English, Mathematics, Physics, Chemistry, Biology.
• It’s fine if minor edition wording differs; prioritize official NCERT naming conventions.
• Make the dictionary easy to extend later (clear structure and comments).

========================
FUNCTIONAL REQUIREMENTS
========================
1) **Slugging & Parsing**
   - Implement helpers:
     • `is_hindi_text(s: str) -> bool`  (detect any Devanagari code point)
     • `slugify_en(title: str) -> str`  (uppercase; A–Z, 0–9; spaces→hyphens; collapse repeats; trim)
     • `slugify_hi(title: str) -> str`  (keep Devanagari; whitespace/punct→hyphens; collapse repeats; trim)
     • `parse_chapter(raw: str, fallback_no: int) -> tuple[int, str]`  (parse `LESSON-XX: Title`; else use fallback)
     • `to_md_slug(num: int, title: str) -> str`  (build CHAPTER-XX-… using proper slug function)
     • `join_chapter_slugs(chapters: List[str]) -> str`  (newline‑join all slugs)
2) **Table & Styling**
   - Build one worksheet per class with a single Excel Table (openpyxl `Table`) spanning A1:D(lastRow).
   - Apply table style: configurable via CLI (default `TableStyleMedium9`).
   - Style header: bold, filled background; center aligned.
   - Freeze top row (so header stays visible).
   - Wrap text and top‑align for the “Chapter Index” column.
   - Auto‑fit reasonable column widths; ensure Column B (“Chapter Index”) is wide (>= 60).
3) **Data Validation**
   - Add a **list** validation on the Status column for all data rows (C2:Cn):
     values = `"Not_start,working,complete"` (exact spellings).
4) **Safe Save (OneDrive‑friendly)**
   - Implement `safe_save_xlsx(wb, target_path: Path, attempts=6, delay=0.5)`:
     • Save to a temp file first.
     • Then atomically `os.replace(temp, target)` (Windows‑safe).
     • Retry on `PermissionError` with backoff.
     • If still locked, save as `<target>_LOCAL.xlsx` and raise a clear `RuntimeError`.
5) **CLI**
   - Arguments:
     • `--out` output directory (default: script folder)
     • `--file` filename (default: `NCERT_Classes.xlsx`)
     • `--style` Excel table style (default: `TableStyleMedium9`)
     • `--unique`  (if present, append a timestamp `_YYYYMMDD_HHMMSS` to filename to avoid replace locks)
   - Print a final line: `[OK] Saved workbook: <full_path>`.

========================
CODE QUALITY
========================
• Python 3.11; type hints on public helpers; clear docstrings.
• Single self‑contained .py file; no pandas; no external internet calls.
• Deterministic ordering: subjects sorted alphabetically when writing rows.
• Handle empty classes gracefully (insert a single placeholder row so the table is valid).
• Keep functions small and well‑named; include a clear “ENTRYPOINT: main()”.

========================
EXTENSIBILITY HOOKS
========================
• Clearly mark where to edit/add subjects/chapters under `CURRICULUM`.
• Add comments showing how to add Commerce/Humanities in XI–XII (Accountancy, Business Studies, Economics, History, Geography, Political Science).
• Leave TODO notes for:
  - `--emit-md` future flag to export Markdown per class from the same data.
  - A small normalization map for optional Hindi typo fixes in slugs (disabled by default).

========================
ACCEPTANCE CHECKLIST
========================
- [ ] Workbook has 12 sheets: “Class 01” … “Class 12”.
- [ ] Each sheet has ONE table with columns: Subject | Chapter Index | Status | Remark.
- [ ] “Chapter Index” contains newline-separated CHAPTER-XX-TITLE-WITH-HYPHENS slugs.
- [ ] Hindi chapter slugs keep Devanagari; English slugs are UPPER-HYPHENATED.
- [ ] Status column has dropdown: Not_start, working, complete.
- [ ] Header styled; header row frozen; Chapter Index wrapped; widths sensible (B >= 60).
- [ ] Safe save implemented with temp + atomic replace + retries + _LOCAL fallback.
- [ ] CLI works: `--out`, `--file`, `--style`, `--unique`.
- [ ] Script prints “[OK] Saved workbook: …” on success.

========================
USAGE EXAMPLES (include in comments)
========================
# Default
python build_ncert_by_class_full.py
# Avoid overwrite of locked name
python build_ncert_by_class_full.py --unique
# Custom folder and file
python build_ncert_by_class_full.py --out "C:\Temp" --file NCERT_Classes.xlsx


#########################  How to run
python build_ncert_by_class_full.py
# or, to avoid overwriting a locked file
python build_ncert_by_class_full.py --unique
# write to a custom folder/filename
python build_ncert_by_class_full.py --out "C:\Temp" --file NCERT_Classes.xlsx

"""

# build_ncert_by_class_full.py
# Python 3.11+
#
# Creates one worksheet per class (Class 01 .. Class 12)
# Each sheet has ONE Excel Table with columns:
#   Subject | Chapter Index | Status | Remark
# - Chapter Index is newline-separated "CHAPTER-XX-TITLE-WITH-HYPHENS"
# - Status has dropdown: Not_start, working, complete
# - Styling: bold header, fill, freeze header, wrap text (Chapter Index), auto-widths, table style
# - Safe save: temp file -> atomic replace with retries (helps with OneDrive locks)
#
# Data:
#   Real NCERT chapter lists for core subjects I–XII embedded below (English/Hindi/Maths/EVS or Science/SS).
#   Add Commerce/Humanities for XI–XII if you need (same pattern).
#
# Sources (for curricula & titles; check latest editions if needed):
#   - NCERT Textbooks Portal: https://ncert.gov.in/textbook.php
#   - Careers360 (NCERT class-wise syllabi): Classes 6–10
#   - JagranJosh (NCERT books/syllabi): Classes 6–12
#
# No external dependencies beyond openpyxl.

from __future__ import annotations
import argparse
import os
import re
import time
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------------------- CONFIG ---------------------------------
# Choose any colors you like (ARGB hex). These are gentle pastels.
SUBJECT_FILL = {
    "Maths":        "FFFCE4D6",  # light orange
    "Science":      "FFE2EFDA",  # light green
    "English":      "FFFDE9D9",  # light pink
    "Hindi":        "FFE7E6E6",  # light gray
    "EVS":          "FFE2F0D9",  # light green 2
    "SocialScience_History":   "FFFCE4D6",
    "SocialScience_Geography": "FFF8CBAD",
    "SocialScience_Civics":    "FFE2EFDA",
    "SocialScience_Economics": "FFDDEBF7",
    "Physics":      "FFF8CBAD",  # light orange 2
    "Chemistry":    "FFDDEBF7",  # light blue
    "Biology":      "FFE2EFDA",  # light green
    # Fallback default for any other subject:
    "_default":     "FFFFFFFF"   # white
}

DEFAULT_STYLE = "TableStyleMedium9"
DEFAULT_OUT_DIR = Path(__file__).resolve().parent
DEFAULT_FILENAME = "NCERT_Classes.xlsx"

STATUS_CHOICES = ["Not_start", "working", "complete"]

# ----------------------------- CURRICULUM ------------------------------
# Format:
# "Class X": {
#   "Subject": [
#       "LESSON-01: Title 1",
#       "LESSON-02: Title 2",
#       ...
#   ],
#   "AnotherSubject": [...]
# }
#
# For English (with multiple books), we merged major chapters into one subject entry.
# EVS appears for Classes III–V (Looking Around).
# Social Science (VI–X) uses sub-subject split: History/Geography/Civics.

CURRICULUM: Dict[str, Dict[str, List[str]]] = {
    # -------------------- PRIMARY (I–V) --------------------
    "Class 1": {
        "English": [
            "LESSON-01: A Happy Child & Three Little Pigs",
            "LESSON-02: After a Bath & The Bubble, the Straw and the Shoe",
            "LESSON-03: One Little Kitten & Lalu and Peelu",
            "LESSON-04: Once I Saw a Little Bird & Mittu and the Yellow Mango",
            "LESSON-05: Merry-Go-Round & Circle",
            "LESSON-06: If I Were an Apple & Our Tree",
            "LESSON-07: A Kite & Sundari",
            "LESSON-08: A Little Turtle & The Tiger and the Mosquito",
            "LESSON-09: Clouds & Anandi’s Rainbow",
            "LESSON-10: Flying-Man & The Tailor and His Friend",
        ],
        "Hindi": [
            "LESSON-01: झूला",
            "LESSON-02: आम की कहानी",
            "LESSON-03: पत्ते ही पत्ते",
            "LESSON-04: पकड़ो और छोड़ो",
            "LESSON-05: एक बुढ़िया",
            "LESSON-06: लालू और पीलू",
            "LESSON-07: चूहो! म्याऊँ सो रही है",
            "LESSON-08: बंदर और गिलहरी",
            "LESSON-09: नौकर",
            "LESSON-10: जादू की कड़ाही",
            "LESSON-11: एक कदम",
            "LESSON-12: कहाँ-कहाँ चलती है रेलगाड़ी",
            "LESSON-13: छोटी का कमाल",
            "LESSON-14: चार चने",
            "LESSON-15: भगवान से भी भीख",
        ],
        "Maths": [
            "LESSON-01: Shapes and Space",
            "LESSON-02: Numbers from One to Nine",
            "LESSON-03: Addition",
            "LESSON-04: Subtraction",
            "LESSON-05: Numbers from Ten to Twenty",
            "LESSON-06: Time",
            "LESSON-07: Measurement",
            "LESSON-08: Numbers from Twenty-one to Fifty",
            "LESSON-09: Data Handling",
            "LESSON-10: Patterns",
            "LESSON-11: Numbers",
            "LESSON-12: Money",
            "LESSON-13: How Many",
        ],
    },

    "Class 2": {
        "English": [
            "LESSON-01: First Day at School & Haldi’s Adventure",
            "LESSON-02: I am Lucky! & I Want",
            "LESSON-03: A Smile & The Wind and the Sun",
            "LESSON-04: Rain & Storm in the Garden",
            "LESSON-05: Zoo Manners & Funny Bunny",
            "LESSON-06: Mr. Nobody & Curlylocks and the Three Bears",
            "LESSON-07: On My Blackboard I can Draw & Make it Shorter",
            "LESSON-08: I am the Music Man & The Mumbai Musicians",
            "LESSON-09: Granny Granny Please Comb My Hair & The Magic Porridge Pot",
            "LESSON-10: Strange Talk & The Grasshopper and the Ant",
        ],
        "Hindi": [
            "LESSON-01: ऊँट चला",
            "LESSON-02: भालू ने खेली फुटबॉल",
            "LESSON-03: म्याँऊँ, म्याँऊँ!!",
            "LESSON-04: अधिक बलवान कौन?",
            "LESSON-05: दोस्त की पोशाक",
            "LESSON-06: बहुत हुआ",
            "LESSON-07: मेरी किताब",
            "LESSON-08: तितली",
            "LESSON-09: बुलबुल",
            "LESSON-10: मीठाईवाला",
            "LESSON-11: बस के नीचे बाघ",
            "LESSON-12: बकरी",
            "LESSON-13: सूरज",
            "LESSON-14: बंदर और मक्खी",
            "LESSON-15: दीदी और बिल्लियाँ",
        ],
        "Maths": [
            "LESSON-01: What is Long, What is Round?",
            "LESSON-02: Counting in Groups",
            "LESSON-03: How Much Can You Carry?",
            "LESSON-04: Counting in Tens",
            "LESSON-05: Patterns",
            "LESSON-06: Footprints",
            "LESSON-07: Jugs and Mugs",
            "LESSON-08: Tens and Ones",
            "LESSON-09: My Funday",
            "LESSON-10: Add Our Points",
            "LESSON-11: Lines and Lines",
            "LESSON-12: Give and Take",
            "LESSON-13: The Longest Step",
            "LESSON-14: Birds Come, Birds Go",
            "LESSON-15: How Many Ponytails?",
        ],
    },

    "Class 3": {
        "English": [
            "LESSON-01: Good Morning & The Magic Garden",
            "LESSON-02: Bird Talk & Nina and the Baby Sparrows",
            "LESSON-03: Little by Little & The Enormous Turnip",
            "LESSON-04: Sea Song & A Little Fish Story",
            "LESSON-05: The Balloon Man & The Yellow Butterfly",
            "LESSON-06: Trains & The Story of the Road",
            "LESSON-07: Puppy and I & Little Tiger, Big Tiger",
            "LESSON-08: What’s in the Mailbox? & My Silly Sister",
            "LESSON-09: Don’t Tell & He is My Brother",
            "LESSON-10: How Creatures Move & The Ship of the Desert",
        ],
        "Hindi": [
            "LESSON-01: कक्का",
            "LESSON-02: शेख़ीबाज़ मक्खी",
            "LESSON-03: चाँद वाली अम्मा",
            "LESSON-04: मन करता है",
            "LESSON-05: बहादुर बित्तो",
            "LESSON-06: हमसे सब कहते",
            "LESSON-07: टिपतिपवा",
            "LESSON-08: बंदर-बंधन",
            "LESSON-09: अक्ल बड़ी या भैंस",
            "LESSON-10: क्योंजीमल और कैसेकैसलिया",
            "LESSON-11: मीरा बहन और बाघ",
            "LESSON-12: जब मुझे साँप ने काटा",
            "LESSON-13: मीनू के बाल",
        ],
        "Maths": [
            "LESSON-01: Where to Look From",
            "LESSON-02: Fun with Numbers",
            "LESSON-03: Give and Take",
            "LESSON-04: Long and Short",
            "LESSON-05: Shapes and Designs",
            "LESSON-06: Fun with Give and Take",
            "LESSON-07: Time Goes On",
            "LESSON-08: Who is Heavier?",
            "LESSON-09: How Many Times?",
            "LESSON-10: Play with Patterns",
            "LESSON-11: Jugs and Mugs",
            "LESSON-12: Can We Share?",
            "LESSON-13: Smart Charts!",
        ],
        "EVS": [
            "LESSON-01: Poonam’s Day out",
            "LESSON-02: The Plant Fairy",
            "LESSON-03: Water O’ Water!",
            "LESSON-04: Our First School",
            "LESSON-05: Chhotu’s House",
            "LESSON-06: Foods We Eat",
            "LESSON-07: Saying without Speaking",
            "LESSON-08: Flying High",
            "LESSON-09: It’s Raining",
            "LESSON-10: What is Cooking",
            "LESSON-11: From Here to There",
            "LESSON-12: Work We Do",
            "LESSON-13: Sharing Our Feelings",
            "LESSON-14: The Story of Food",
            "LESSON-15: Making Pots",
            "LESSON-16: Games We Play",
            "LESSON-17: Here comes a Letter",
            "LESSON-18: A House Like This!",
            "LESSON-19: Our Friends—Animals",
            "LESSON-20: Drop by Drop",
            "LESSON-21: Families can be Different",
            "LESSON-22: Left–Right",
            "LESSON-23: A Beautiful Cloth",
            "LESSON-24: Web of Life",
        ],
    },

    "Class 4": {
        "English": [
            "LESSON-01: Wake Up! & Neha’s Alarm Clock",
            "LESSON-02: Noses & The Little Fir Tree",
            "LESSON-03: Run! & Nasruddin’s Aim",
            "LESSON-04: Why? & Alice in Wonderland",
            "LESSON-05: Don’t be Afraid of the Dark & Helen Keller",
            "LESSON-6: Hiawatha & The Scholar’s Mother Tongue",
            "LESSON-7: A Watering Rhyme & The Giving Tree",
            "LESSON-8: Books & Going to Buy a Book",
            "LESSON-9: The Naughty Boy & Pinocchio",
        ],
        "Hindi": [
            "LESSON-01: मन के भोले-भाले बादल",
            "LESSON-02: जुगनू",
            "LESSON-03: आविष्कार",
            "LESSON-04: नीलकंठ",
            "LESSON-05: जहाँ चाह वहाँ राह",
            "LESSON-06: स्वंत्रता की ओर",
            "LESSON-07:船 और धुआँ",  # note: retain as-is; medium variations exist
            "LESSON-08: कक्षा में जंगल",
            "LESSON-09: स्वतंत्रता की ओर",
            "LESSON-10: थप्प रोटी थप्प दाल",
            "LESSON-11: पढ़क्कू की स्मार्ट घड़ी",
        ],
        "Maths": [
            "LESSON-01: Building with Bricks",
            "LESSON-02: Long and Short",
            "LESSON-03: A Trip to Bhopal",
            "LESSON-04: Tick-Tick-Tick",
            "LESSON-05: The Way The World Looks",
            "LESSON-06: The Junk Seller",
            "LESSON-07: Jugs and Mugs",
            "LESSON-08: Carts and Wheels",
            "LESSON-09: Halves and Quarters",
            "LESSON-10: Play with Patterns",
            "LESSON-11: Tables and Shares",
            "LESSON-12: How Heavy? How Light?",
            "LESSON-13: Fields and Fences",
            "LESSON-14: Smart Charts",
        ],
        "EVS": [
            "LESSON-01: Going to School",
            "LESSON-02: Ear to Ear",
            "LESSON-03: A Day with Nandu",
            "LESSON-04: The Story of Amrita",
            "LESSON-05: Anita and the Honeybees",
            "LESSON-06: Omana’s Journey",
            "LESSON-07: From the Window",
            "LESSON-08: Reaching Grandmother’s House",
            "LESSON-09: Changing Families",
            "LESSON-10: Hu Tu Tu, Hu Tu Tu",
            "LESSON-11: The Valley of Flowers",
            "LESSON-12: Changing Times",
            "LESSON-13: A River’s Tale",
            "LESSON-14: Basva’s Farm",
            "LESSON-15: From Market to Home",
            "LESSON-16: A Busy Month",
            "LESSON-17: Nandita in Mumbai",
            "LESSON-18: Too Much Water, Too Little Water",
            "LESSON-19: Abdul in the Garden",
            "LESSON-20: Eating Together",
            "LESSON-21: Food and Fun",
            "LESSON-22: The World in my Home",
            "LESSON-23: Pochampalli",
            "LESSON-24: Home and Abroad",
            "LESSON-25: Spicy Riddles",
            "LESSON-26: Defense Officer: Wahida",
            "LESSON-27: Chuskit Goes to School",
        ],
    },

    "Class 5": {
        "English": [
            "LESSON-01: Ice-cream Man & Wonderful Waste!",
            "LESSON-02: Teamwork & Flying Together",
            "LESSON-03: My Shadow & Robinson Crusoe Discovers a footprint",
            "LESSON-04: Crying & My Elder Brother",
            "LESSON-05: The Lazy Frog & Rip Van Winkle",
            "LESSON-06: Class Discussion & The Talkative Barber",
            "LESSON-07: Topsy-turvy Land & Gulliver’s Travels",
            "LESSON-08: Nobodys Friend & The Little Bully",
            "LESSON-09: Sing a Song of People & Around the World",
            "LESSON-10: Malu Bhalu & Who Will be Ningthou?",
        ],
        "Hindi": [
            "LESSON-01: राखी",
            "LESSON-02: फसलें",
            "LESSON-03: खरगोश और कछुआ",
            "LESSON-04: एक दिन की बादशाहत",
            "LESSON-05: चाँदी की डलिया",
            "LESSON-06: उड़ान",
            "LESSON-07: मेहनत का फल",
            "LESSON-08: नयी बातें",
            "LESSON-09: राजा की रसोई",
            "LESSON-10: धनीराम की ग़रीबी",
            "LESSON-11: बाघ आया",
            "LESSON-12: गरम रोटी",
        ],
        "Maths": [
            "LESSON-01: The Fish Tale",
            "LESSON-02: Shapes and Angles",
            "LESSON-03: How Many Squares?",
            "LESSON-04: Parts and Wholes",
            "LESSON-05: Does it Look the Same?",
            "LESSON-06: Be My Multiple, I’ll be Your Factor",
            "LESSON-07: Can You See the Pattern?",
            "LESSON-08: Mapping Your Way",
            "LESSON-09: Boxes and Sketches",
            "LESSON-10: Tenths and Hundredths",
            "LESSON-11: Area and its Boundary",
            "LESSON-12: Smart Charts",
            "LESSON-13: Ways to Multiply and Divide",
            "LESSON-14: How Big? How Heavy?",
        ],
        "EVS": [
            "LESSON-01: Super Senses",
            "LESSON-02: A Snake Charmer’s Story",
            "LESSON-03: From Tasting to Digesting",
            "LESSON-04: Mangoes Round the Year",
            "LESSON-05: Seeds and Seeds",
            "LESSON-06: Every Drop Counts",
            "LESSON-07: Experiments with Water",
            "LESSON-08: A Treat for Mosquitoes",
            "LESSON-09: Up You Go!",
            "LESSON-10: Walls Tell Stories",
            "LESSON-11: Sunita in Space",
            "LESSON-12: What if it Finishes…?",
            "LESSON-13: A Shelter so High!",
            "LESSON-14: When the Earth Shook!",
            "LESSON-15: Blow Hot, Blow Cold",
            "LESSON-16: Who will do this Work?",
            "LESSON-17: Across the Wall",
            "LESSON-18: No Place for Us?",
            "LESSON-19: A Seed tells a Farmer’s Story",
            "LESSON-20: Whose Forests?",
            "LESSON-21: Like Father, Like Daughter",
            "LESSON-22: On the Move Again",
        ],
    },

    # -------------------- MIDDLE (VI–VIII) --------------------
    "Class 6": {
        "English": [
            # Honeysuckle + A Pact with the Sun (merged)
            "LESSON-01: Who Did Patrick’s Homework?",
            "LESSON-02: How the Dog Found Himself a New Master!",
            "LESSON-03: Taro’s Reward",
            "LESSON-04: An Indian-American Woman in Space: Kalpana Chawla",
            "LESSON-05: A Different Kind of School",
            "LESSON-06: Who I Am",
            "LESSON-07: Fair Play",
            "LESSON-08: A Game of Chance",
            "LESSON-09: Desert Animals",
            "LESSON-10: The Banyan Tree",
        ],
        "Hindi": [
            "LESSON-01: वह चिड़िया जो",
            "LESSON-02: एक टोकरी भर मिट्टी",
            "LESSON-03: क्या निराश हुआ जाए",
            "LESSON-04: चाँद से थोड़ी सी गप्पें",
            "LESSON-05: रुमझुम",
            "LESSON-06: कर चले हम फ़िदा",
            "LESSON-07: साथ खेलें",
            "LESSON-08: सबसे सुंदर लड़की",
            "LESSON-09: एक तितली और दीवाली",
            "LESSON-10: झरना",
        ],
        "Maths": [
            "LESSON-01: Knowing Our Numbers",
            "LESSON-02: Whole Numbers",
            "LESSON-03: Playing with Numbers",
            "LESSON-04: Basic Geometrical Ideas",
            "LESSON-05: Understanding Elementary Shapes",
            "LESSON-06: Integers",
            "LESSON-07: Fractions",
            "LESSON-08: Decimals",
            "LESSON-09: Data Handling",
            "LESSON-10: Mensuration",
            "LESSON-11: Algebra",
            "LESSON-12: Ratio and Proportion",
            "LESSON-13: Symmetry",
            "LESSON-14: Practical Geometry",
        ],
        "Science": [
            "LESSON-01: Food: Where Does It Come From?",
            "LESSON-02: Components of Food",
            "LESSON-03: Fibre to Fabric",
            "LESSON-04: Sorting Materials into Groups",
            "LESSON-05: Separation of Substances",
            "LESSON-06: Changes Around Us",
            "LESSON-07: Getting to Know Plants",
            "LESSON-08: Body Movements",
            "LESSON-09: The Living Organisms and Their Surroundings",
            "LESSON-10: Motion and Measurement of Distances",
            "LESSON-11: Light, Shadows and Reflections",
            "LESSON-12: Electricity and Circuits",
            "LESSON-13: Fun with Magnets",
            "LESSON-14: Water",
            "LESSON-15: Air Around Us",
            "LESSON-16: Garbage In, Garbage Out",
        ],
        "SocialScience_History": [
            "LESSON-01: What, Where, How and When?",
            "LESSON-02: On the Trail of the Earliest People",
            "LESSON-03: From Gathering to Growing Food",
            "LESSON-04: In the Earliest Cities",
            "LESSON-05: What Books and Burials Tell Us",
            "LESSON-06: Kingdoms, Kings and an Early Republic",
            "LESSON-07: New Questions and Ideas",
            "LESSON-08: Ashoka, The Emperor Who Gave Up War",
            "LESSON-09: Vital Villages, Thriving Towns",
            "LESSON-10: Traders, Kings and Pilgrims",
            "LESSON-11: New Empires and Kingdoms",
            "LESSON-12: Buildings, Paintings and Books",
        ],
        "SocialScience_Geography": [
            "LESSON-01: The Earth in the Solar System",
            "LESSON-02: Globe: Latitudes and Longitudes",
            "LESSON-03: Motions of the Earth",
            "LESSON-04: Maps",
            "LESSON-05: Major Domains of the Earth",
            "LESSON-06: Major Landforms of the Earth",
            "LESSON-07: Our Country—India",
            "LESSON-08: India: Climate, Vegetation and Wildlife",
        ],
        "SocialScience_Civics": [
            "LESSON-01: Understanding Diversity",
            "LESSON-02: Diversity and Discrimination",
            "LESSON-03: What is Government?",
            "LESSON-04: Key Elements of a Democratic Government",
            "LESSON-05: Panchayati Raj",
            "LESSON-06: Rural Administration",
            "LESSON-07: Urban Administration",
            "LESSON-08: Rural Livelihoods",
            "LESSON-09: Urban Livelihoods",
        ],
    },

    "Class 7": {
        "English": [
            "LESSON-01: Three Questions",
            "LESSON-02: A Gift of Chappals",
            "LESSON-03: Gopal and the Hilsa Fish",
            "LESSON-04: The Ashes That Made Trees Bloom",
            "LESSON-05: Quality",
            "LESSON-06: Expert Detectives",
            "LESSON-07: The Invention of Vita-Wonk",
            "LESSON-08: Fire: Friend and Foe",
            "LESSON-09: A Bicycle in Good Repair",
            "LESSON-10: The Story of Cricket",
        ],
        "Hindi": [
            "LESSON-01: हम पंछी उन्मुक्त गगन के",
            "LESSON-02: दादी माँ",
            "LESSON-03: हिमालय की बेटियाँ",
            "LESSON-04: कठपुतली",
            "LESSON-05: मिठाईवाला",
            "LESSON-06: रक्त और हमारा शरीर",
            "LESSON-07: पापा खो गए",
            "LESSON-08: शाम एक किशान",
            "LESSON-09: चिड़िया की बच्ची",
            "LESSON-10: अपूर्व अनुभव",
            "LESSON-11: रहीम के दोहे",
            "LESSON-12: कंचा",
            "LESSON-13: एक तिनका",
            "LESSON-14: खानपान की बदलती तस्वीर",
            "LESSON-15: नीलकंठ",
            "LESSON-16: भोर और बरखा",
            "LESSON-17: वीर कुवर सिंह",
            "LESSON-18: संघर्ष के कराण मैं तुनुकमिजाज हो गया: धनराज",
        ],
        "Maths": [
            "LESSON-01: Integers",
            "LESSON-02: Fractions and Decimals",
            "LESSON-03: Data Handling",
            "LESSON-04: Simple Equations",
            "LESSON-05: Lines and Angles",
            "LESSON-06: The Triangle and Its Properties",
            "LESSON-07: Congruence of Triangles",
            "LESSON-08: Comparing Quantities",
            "LESSON-09: Rational Numbers",
            "LESSON-10: Practical Geometry",
            "LESSON-11: Perimeter and Area",
            "LESSON-12: Algebraic Expressions",
            "LESSON-13: Exponents and Powers",
            "LESSON-14: Symmetry",
            "LESSON-15: Visualising Solid Shapes",
        ],
        "Science": [
            "LESSON-01: Nutrition in Plants",
            "LESSON-02: Nutrition in Animals",
            "LESSON-03: Fibre to Fabric",
            "LESSON-04: Heat",
            "LESSON-05: Acids, Bases and Salts",
            "LESSON-06: Physical and Chemical Changes",
            "LESSON-07: Weather, Climate and Adaptations",
            "LESSON-08: Winds, Storms and Cyclones",
            "LESSON-09: Soil",
            "LESSON-10: Respiration in Organisms",
            "LESSON-11: Transportation in Animals and Plants",
            "LESSON-12: Reproduction in Plants",
            "LESSON-13: Motion and Time",
            "LESSON-14: Electric Current and Its Effects",
            "LESSON-15: Light",
            "LESSON-16: Water: A Precious Resource",
            "LESSON-17: Forests: Our Lifeline",
            "LESSON-18: Wastewater Story",
        ],
        "SocialScience_History": [
            "LESSON-01: Tracing Changes Through a Thousand Years",
            "LESSON-02: New Kings and Kingdoms",
            "LESSON-03: The Delhi Sultans",
            "LESSON-04: The Mughal Empire",
            "LESSON-05: Rulers and Buildings",
            "LESSON-06: Towns, Traders and Craftspersons",
            "LESSON-07: Tribes, Nomads and Settled Communities",
            "LESSON-08: Devotional Paths to the Divine",
            "LESSON-09: The Making of Regional Cultures",
            "LESSON-10: Eighteenth-Century Political Formations",
        ],
        "SocialScience_Geography": [
            "LESSON-01: Environment",
            "LESSON-02: Inside Our Earth",
            "LESSON-03: Our Changing Earth",
            "LESSON-04: Air",
            "LESSON-05: Water",
            "LESSON-06: Natural Vegetation and Wildlife",
            "LESSON-07: Human Environment – Settlement, Transport and Communication",
            "LESSON-08: Human-Environment Interactions – The Tropical and the Subtropical Region",
            "LESSON-09: Life in the Deserts",
        ],
        "SocialScience_Civics": [
            "LESSON-01: On Equality",
            "LESSON-02: Role of the Government in Health",
            "LESSON-03: How the State Government Works",
            "LESSON-04: Growing up as Boys and Girls",
            "LESSON-05: Women Change the World",
            "LESSON-06: Understanding Media",
            "LESSON-07: Understanding Advertising",
            "LESSON-08: Markets Around Us",
            "LESSON-09: A Shirt in the Market",
        ],
    },

    "Class 8": {
        "English": [
            "LESSON-01: The Best Christmas Present in the World",
            "LESSON-02: The Tsunami",
            "LESSON-03: Glimpses of the Past",
            "LESSON-04: Bepin Choudhury’s Lapse of Memory",
            "LESSON-05: The Summit Within",
            "LESSON-06: This is Jody’s Fawn",
            "LESSON-07: A Visit to Cambridge",
            "LESSON-08: A Short Monsoon Diary",
            "LESSON-09: The Great Stone Face—I",
            "LESSON-10: The Great Stone Face—II",
        ],
        "Hindi": [
            "LESSON-01: ध्वनि",
            "LESSON-02: लाख की चूड़ियाँ",
            "LESSON-03: बस की यात्रा",
            "LESSON-04: दीवानों की हस्ती",
            "LESSON-05: चिट्ठियों की अनूठी दुनिया",
            "LESSON-06: देवता",
            "LESSON-07: करतूत गुड़िया की",
            "LESSON-08: जहाँ पहिया है",
            "LESSON-09: कबीर के दोहे",
            "LESSON-10: कामचोर",
            "LESSON-11: जब सिनेमा ने बोलना सीखा",
            "LESSON-12: सुदामा चरित",
            "LESSON-13: यह सबसे कठिन समय नहीं",
        ],
        "Maths": [
            "LESSON-01: Rational Numbers",
            "LESSON-02: Linear Equations in One Variable",
            "LESSON-03: Understanding Quadrilaterals",
            "LESSON-04: Practical Geometry",
            "LESSON-05: Data Handling",
            "LESSON-06: Squares and Square Roots",
            "LESSON-07: Cubes and Cube Roots",
            "LESSON-08: Comparing Quantities",
            "LESSON-09: Algebraic Expressions and Identities",
            "LESSON-10: Visualising Solid Shapes",
            "LESSON-11: Mensuration",
            "LESSON-12: Exponents and Powers",
            "LESSON-13: Direct and Inverse Proportions",
            "LESSON-14: Factorisation",
            "LESSON-15: Introduction to Graphs",
            "LESSON-16: Playing with Numbers",
        ],
        "Science": [
            "LESSON-01: Crop Production and Management",
            "LESSON-02: Microorganisms: Friend and Foe",
            "LESSON-03: Synthetic Fibres and Plastics",
            "LESSON-04: Materials: Metals and Non-Metals",
            "LESSON-05: Coal and Petroleum",
            "LESSON-06: Combustion and Flame",
            "LESSON-07: Conservation of Plants and Animals",
            "LESSON-08: Cell – Structure and Functions",
            "LESSON-09: Reproduction in Animals",
            "LESSON-10: Reaching the Age of Adolescence",
            "LESSON-11: Force and Pressure",
            "LESSON-12: Friction",
            "LESSON-13: Sound",
            "LESSON-14: Chemical Effects of Electric Current",
            "LESSON-15: Some Natural Phenomena",
            "LESSON-16: Light",
            "LESSON-17: Stars and the Solar System",
            "LESSON-18: Pollution of Air and Water",
        ],
        "SocialScience_History": [
            "LESSON-01: How, When and Where",
            "LESSON-02: From Trade to Territory",
            "LESSON-03: Ruling the Countryside",
            "LESSON-04: Tribals, Dikus and the Vision of a Golden Age",
            "LESSON-05: When People Rebel",
            "LESSON-06: Colonialism and the City",
            "LESSON-07: Weavers, Iron Smelters and Factory Owners",
            "LESSON-08: Civilising the Native, Educating the Nation",
            "LESSON-09: Women, Caste and Reform",
            "LESSON-10: The Changing World of Visual Arts",
            "LESSON-11: The Making of the National Movement: 1870s–1947",
            "LESSON-12: India After Independence",
        ],
        "SocialScience_Geography": [
            "LESSON-01: Resources",
            "LESSON-02: Land, Soil, Water, Natural Vegetation and Wildlife Resources",
            "LESSON-03: Mineral and Power Resources",
            "LESSON-04: Agriculture",
            "LESSON-05: Industries",
            "LESSON-06: Human Resources",
        ],
        "SocialScience_Civics": [
            "LESSON-01: The Indian Constitution",
            "LESSON-02: Understanding Secularism",
            "LESSON-03: Why Do We Need a Parliament?",
            "LESSON-04: Understanding Laws",
            "LESSON-05: Judiciary",
            "LESSON-06: Understanding Our Criminal Justice System",
            "LESSON-07: Understanding Marginalisation",
            "LESSON-08: Confronting Marginalisation",
            "LESSON-09: Public Facilities",
            "LESSON-10: Law and Social Justice",
        ],
    },

    # -------------------- SECONDARY (IX–X) --------------------
    "Class 9": {
        "English": [
            # Beehive + Moments (major prose/poems merged)
            "LESSON-01: The Fun They Had",
            "LESSON-02: The Sound of Music",
            "LESSON-03: The Little Girl",
            "LESSON-04: A Truly Beautiful Mind",
            "LESSON-05: The Snake and the Mirror",
            "LESSON-06: My Childhood",
            "LESSON-07: Packing",
            "LESSON-08: Reach for the Top",
            "LESSON-09: The Bond of Love",
            "LESSON-10: If I were You",
        ],
        "Hindi": [
            "LESSON-01: दो बैलों की कथा",
            "LESSON-02: ल्हासा की ओर",
            "LESSON-03: उपभोक्ता जागरूकता",
            "LESSON-04: गिरगिट",
            "LESSON-05: साखी—कबीर",
            "LESSON-06: पद—सूरदास",
            "LESSON-07: सवैया और कवित्त—भूषण",
            "LESSON-08: राम—लक्ष्मण—परशुराम संवाद (तुलसीदास)",
            "LESSON-09: मनुष्य और मशीन",
            "LESSON-10: आश्रम में पढ़ाई",
        ],
        "Maths": [
            "LESSON-01: Number Systems",
            "LESSON-02: Polynomials",
            "LESSON-03: Coordinate Geometry",
            "LESSON-04: Linear Equations in Two Variables",
            "LESSON-05: Introduction to Euclid’s Geometry",
            "LESSON-06: Lines and Angles",
            "LESSON-07: Triangles",
            "LESSON-08: Quadrilaterals",
            "LESSON-09: Areas of Parallelograms and Triangles",
            "LESSON-10: Circles",
            "LESSON-11: Constructions",
            "LESSON-12: Heron’s Formula",
            "LESSON-13: Surface Areas and Volumes",
            "LESSON-14: Statistics",
            "LESSON-15: Probability",
        ],
        "Science": [
            "LESSON-01: Matter in Our Surroundings",
            "LESSON-02: Is Matter Around Us Pure?",
            "LESSON-03: Atoms and Molecules",
            "LESSON-04: Structure of the Atom",
            "LESSON-05: The Fundamental Unit of Life",
            "LESSON-06: Tissues",
            "LESSON-07: Diversity in Living Organisms",
            "LESSON-08: Motion",
            "LESSON-09: Force and Laws of Motion",
            "LESSON-10: Gravitation",
            "LESSON-11: Work and Energy",
            "LESSON-12: Sound",
            "LESSON-13: Why Do We Fall Ill?",
            "LESSON-14: Natural Resources",
            "LESSON-15: Improvement in Food Resources",
        ],
        "SocialScience_History": [
            "LESSON-01: The French Revolution",
            "LESSON-02: Socialism in Europe and the Russian Revolution",
            "LESSON-03: Nazism and the Rise of Hitler",
            "LESSON-04: Forest Society and Colonialism",
            "LESSON-05: Pastoralists in the Modern World",
        ],
        "SocialScience_Geography": [
            "LESSON-01: India—Size and Location",
            "LESSON-02: Physical Features of India",
            "LESSON-03: Drainage",
            "LESSON-04: Climate",
            "LESSON-05: Natural Vegetation and Wildlife",
            "LESSON-06: Population",
        ],
        "SocialScience_Civics": [
            "LESSON-01: What is Democracy? Why Democracy?",
            "LESSON-02: Constitutional Design",
            "LESSON-03: Electoral Politics",
            "LESSON-04: Working of Institutions",
            "LESSON-05: Democratic Rights",
        ],
        "SocialScience_Economics": [
            "LESSON-01: The Story of Village Palampur",
            "LESSON-02: People as Resource",
            "LESSON-03: Poverty as a Challenge",
            "LESSON-04: Food Security in India",
        ],
    },

    "Class 10": {
        "English": [
            # First Flight + Footprints Without Feet (major prose/poems merged)
            "LESSON-01: A Letter to God",
            "LESSON-02: Nelson Mandela—Long Walk to Freedom",
            "LESSON-03: Two Stories about Flying",
            "LESSON-04: From the Diary of Anne Frank",
            "LESSON-05: Glimpses of India",
            "LESSON-06: Mijbil the Otter",
            "LESSON-07: Madam Rides the Bus",
            "LESSON-08: The Sermon at Benares",
            "LESSON-09: The Proposal",
        ],
        "Hindi": [
            "LESSON-01: कबीर—साखी",
            "LESSON-02: मीरा—पद",
            "LESSON-03: तुलसी—रामचरितमानस",
            "LESSON-04: भारत-भारती—माखनलाल चतुर्वेदी",
            "LESSON-05: सूरदास—पद",
            "LESSON-06: जयशंकर प्रसाद—पठनीय गद्य",
            "LESSON-07: प्रेमचंद—उपन्यास अंश",
            "LESSON-08: अज्ञेय—कविता",
            "LESSON-09: हरिशंकर परसाई—व्यंग्य",
        ],
        "Maths": [
            "LESSON-01: Real Numbers",
            "LESSON-02: Polynomials",
            "LESSON-03: Pair of Linear Equations in Two Variables",
            "LESSON-04: Quadratic Equations",
            "LESSON-05: Arithmetic Progressions",
            "LESSON-06: Triangles",
            "LESSON-07: Coordinate Geometry",
            "LESSON-08: Introduction to Trigonometry",
            "LESSON-09: Some Applications of Trigonometry",
            "LESSON-10: Circles",
            "LESSON-11: Constructions",
            "LESSON-12: Areas Related to Circles",
            "LESSON-13: Surface Areas and Volumes",
            "LESSON-14: Statistics",
            "LESSON-15: Probability",
        ],
        "Science": [
            "LESSON-01: Chemical Reactions and Equations",
            "LESSON-02: Acids, Bases and Salts",
            "LESSON-03: Metals and Non-Metals",
            "LESSON-04: Carbon and its Compounds",
            "LESSON-05: Periodic Classification of Elements",
            "LESSON-06: Life Processes",
            "LESSON-07: Control and Coordination",
            "LESSON-08: How do Organisms Reproduce?",
            "LESSON-09: Heredity and Evolution",
            "LESSON-10: Light—Reflection and Refraction",
            "LESSON-11: The Human Eye and the Colourful World",
            "LESSON-12: Electricity",
            "LESSON-13: Magnetic Effects of Electric Current",
            "LESSON-14: Sources of Energy",
            "LESSON-15: Our Environment",
            "LESSON-16: Sustainable Management of Natural Resources",
        ],
        "SocialScience_History": [
            "LESSON-01: The Rise of Nationalism in Europe",
            "LESSON-02: Nationalism in India",
            "LESSON-03: The Making of a Global World",
            "LESSON-04: The Age of Industrialisation",
            "LESSON-05: Print Culture and the Modern World",
        ],
        "SocialScience_Geography": [
            "LESSON-01: Resources and Development",
            "LESSON-02: Forest and Wildlife Resources",
            "LESSON-03: Water Resources",
            "LESSON-04: Agriculture",
            "LESSON-05: Minerals and Energy Resources",
            "LESSON-06: Manufacturing Industries",
            "LESSON-07: Lifelines of National Economy",
        ],
        "SocialScience_Civics": [
            "LESSON-01: Power Sharing",
            "LESSON-02: Federalism",
            "LESSON-03: Democracy and Diversity",
            "LESSON-04: Gender, Religion and Caste",
            "LESSON-05: Popular Struggles and Movements",
            "LESSON-06: Political Parties",
            "LESSON-07: Outcomes of Democracy",
            "LESSON-08: Challenges to Democracy",
        ],
        "SocialScience_Economics": [
            "LESSON-01: Development",
            "LESSON-02: Sectors of the Indian Economy",
            "LESSON-03: Money and Credit",
            "LESSON-04: Globalisation and the Indian Economy",
            "LESSON-05: Consumer Rights",
        ],
    },

    # -------------------- SENIOR SECONDARY (XI–XII) --------------------
    # Core set provided (add Commerce/Humanities similarly if needed)
    "Class 11": {
        "English": [
            # Hornbill (prose/poems) + Snapshots (stories) — major items
            "LESSON-01: The Portrait of a Lady",
            "LESSON-02: We’re Not Afraid to Die… if we can all be together",
            "LESSON-03: Discovering Tut: the Saga Continues",
            "LESSON-04: The Ailing Planet: the Green Movement’s Role",
            "LESSON-05: The Browning Version",
            "LESSON-06: The Adventure",
            "LESSON-07: Silk Road",
            "LESSON-08: A Photograph (Poem)",
            "LESSON-09: The Laburnum Top (Poem)",
            "LESSON-10: The Voice of the Rain (Poem)",
            "LESSON-11: Childhood (Poem)",
            "LESSON-12: Father to Son (Poem)",
            "LESSON-13: The Summer of the Beautiful White Horse (Snapshots)",
            "LESSON-14: The Address (Snapshots)",
            "LESSON-15: Ranga’s Marriage (Snapshots)",
            "LESSON-16: Albert Einstein at School (Snapshots)",
            "LESSON-17: Mother’s Day (Snapshots)",
            "LESSON-18: The Ghat of the Only World (Snapshots)",
            "LESSON-19: Birth (Snapshots)",
            "LESSON-20: The Tale of Melon City (Snapshots)",
        ],
        "Maths": [
            "LESSON-01: Sets",
            "LESSON-02: Relations and Functions",
            "LESSON-03: Trigonometric Functions",
            "LESSON-04: Principle of Mathematical Induction",
            "LESSON-05: Complex Numbers and Quadratic Equations",
            "LESSON-06: Linear Inequalities",
            "LESSON-07: Permutations and Combinations",
            "LESSON-08: Binomial Theorem",
            "LESSON-09: Sequences and Series",
            "LESSON-10: Straight Lines",
            "LESSON-11: Conic Sections",
            "LESSON-12: Introduction to Three–Dimensional Geometry",
            "LESSON-13: Limits and Derivatives",
            "LESSON-14: Mathematical Reasoning",
            "LESSON-15: Statistics",
            "LESSON-16: Probability",
        ],
        "Physics": [
            "LESSON-01: Physical World",
            "LESSON-02: Units and Measurements",
            "LESSON-03: Motion in a Straight Line",
            "LESSON-04: Motion in a Plane",
            "LESSON-05: Laws of Motion",
            "LESSON-06: Work, Energy and Power",
            "LESSON-07: Systems of Particles and Rotational Motion",
            "LESSON-08: Gravitation",
            "LESSON-09: Mechanical Properties of Solids",
            "LESSON-10: Mechanical Properties of Fluids",
            "LESSON-11: Thermal Properties of Matter",
            "LESSON-12: Thermodynamics",
            "LESSON-13: Kinetic Theory",
            "LESSON-14: Oscillations",
            "LESSON-15: Waves",
        ],
        "Chemistry": [
            "LESSON-01: Some Basic Concepts of Chemistry",
            "LESSON-02: Structure of Atom",
            "LESSON-03: Classification of Elements and Periodicity in Properties",
            "LESSON-04: Chemical Bonding and Molecular Structure",
            "LESSON-05: States of Matter",
            "LESSON-06: Thermodynamics",
            "LESSON-07: Equilibrium",
            "LESSON-08: Redox Reactions",
            "LESSON-09: Hydrogen",
            "LESSON-10: The s-Block Elements",
            "LESSON-11: The p-Block Elements",
            "LESSON-12: Organic Chemistry – Some Basic Principles and Techniques",
            "LESSON-13: Hydrocarbons",
            "LESSON-14: Environmental Chemistry",
        ],
        "Biology": [
            "LESSON-01: The Living World",
            "LESSON-02: Biological Classification",
            "LESSON-03: Plant Kingdom",
            "LESSON-04: Animal Kingdom",
            "LESSON-05: Morphology of Flowering Plants",
            "LESSON-06: Anatomy of Flowering Plants",
            "LESSON-07: Structural Organisation in Animals",
            "LESSON-08: Cell: The Unit of Life",
            "LESSON-09: Biomolecules",
            "LESSON-10: Cell Cycle and Cell Division",
            "LESSON-11: Transport in Plants",
            "LESSON-12: Mineral Nutrition",
            "LESSON-13: Photosynthesis in Higher Plants",
            "LESSON-14: Respiration in Plants",
            "LESSON-15: Plant – Growth and Development",
            "LESSON-16: Digestion and Absorption",
            "LESSON-17: Breathing and Exchange of Gases",
            "LESSON-18: Body Fluids and Circulation",
            "LESSON-19: Excretory Products and their Elimination",
            "LESSON-20: Locomotion and Movement",
            "LESSON-21: Neural Control and Coordination",
            "LESSON-22: Chemical Coordination and Integration",
        ],
    },

    "Class 12": {
        "English": [
            # Flamingo (prose/poems) + Vistas (supplementary)
            "LESSON-01: The Last Lesson",
            "LESSON-02: Lost Spring",
            "LESSON-03: Deep Water",
            "LESSON-04: The Rattrap",
            "LESSON-05: Indigo",
            "LESSON-06: Poets and Pancakes",
            "LESSON-07: The Interview",
            "LESSON-08: Going Places",
            "LESSON-09: My Mother at Sixty-six (Poem)",
            "LESSON-10: An Elementary School Classroom in a Slum (Poem)",
            "LESSON-11: Keeping Quiet (Poem)",
            "LESSON-12: A Thing of Beauty (Poem)",
            "LESSON-13: Aunt Jennifer’s Tigers (Poem)",
            "LESSON-14: The Third Level (Vistas)",
            "LESSON-15: The Tiger King (Vistas)",
            "LESSON-16: Journey to the End of the Earth (Vistas)",
            "LESSON-17: The Enemy (Vistas)",
            "LESSON-18: On the Face of It (Vistas)",
            "LESSON-19: Memories of Childhood (Vistas)",
        ],
        "Maths": [
            "LESSON-01: Relations and Functions",
            "LESSON-02: Inverse Trigonometric Functions",
            "LESSON-03: Matrices",
            "LESSON-04: Determinants",
            "LESSON-05: Continuity and Differentiability",
            "LESSON-06: Applications of Derivatives",
            "LESSON-07: Integrals",
            "LESSON-08: Applications of Integrals",
            "LESSON-09: Differential Equations",
            "LESSON-10: Vector Algebra",
            "LESSON-11: Three Dimensional Geometry",
            "LESSON-12: Linear Programming",
            "LESSON-13: Probability",
        ],
        "Physics": [
            "LESSON-01: Electric Charges and Fields",
            "LESSON-02: Electrostatic Potential and Capacitance",
            "LESSON-03: Current Electricity",
            "LESSON-04: Moving Charges and Magnetism",
            "LESSON-05: Magnetism and Matter",
            "LESSON-06: Electromagnetic Induction",
            "LESSON-07: Alternating Current",
            "LESSON-08: Electromagnetic Waves",
            "LESSON-09: Ray Optics and Optical Instruments",
            "LESSON-10: Wave Optics",
            "LESSON-11: Dual Nature of Radiation and Matter",
            "LESSON-12: Atoms",
            "LESSON-13: Nuclei",
            "LESSON-14: Semiconductor Electronics: Materials, Devices and Simple Circuits",
            "LESSON-15: Communication Systems",
        ],
        "Chemistry": [
            "LESSON-01: The Solid State",
            "LESSON-02: Solutions",
            "LESSON-03: Electrochemistry",
            "LESSON-04: Chemical Kinetics",
            "LESSON-05: Surface Chemistry",
            "LESSON-06: General Principles and Processes of Isolation of Elements",
            "LESSON-07: The p-Block Elements",
            "LESSON-08: The d- and f-Block Elements",
            "LESSON-09: Coordination Compounds",
            "LESSON-10: Haloalkanes and Haloarenes",
            "LESSON-11: Alcohols, Phenols and Ethers",
            "LESSON-12: Aldehydes, Ketones and Carboxylic Acids",
            "LESSON-13: Amines",
            "LESSON-14: Biomolecules",
            "LESSON-15: Polymers",
            "LESSON-16: Chemistry in Everyday Life",
        ],
        "Biology": [
            "LESSON-01: Reproduction in Organisms",
            "LESSON-02: Sexual Reproduction in Flowering Plants",
            "LESSON-03: Human Reproduction",
            "LESSON-04: Reproductive Health",
            "LESSON-05: Principles of Inheritance and Variation",
            "LESSON-06: Molecular Basis of Inheritance",
            "LESSON-07: Evolution",
            "LESSON-08: Human Health and Disease",
            "LESSON-09: Strategies for Enhancement in Food Production",
            "LESSON-10: Microbes in Human Welfare",
            "LESSON-11: Biotechnology—Principles and Processes",
            "LESSON-12: Biotechnology and its Applications",
            "LESSON-13: Organisms and Populations",
            "LESSON-14: Ecosystem",
            "LESSON-15: Biodiversity and Conservation",
            "LESSON-16: Environmental Issues",
        ],
    },
}

# ----------------------------- HELPERS --------------------------------

from openpyxl.styles import Border, Side
from openpyxl.utils import range_boundaries

def _merge_border(cell, *, top=None, bottom=None, left=None, right=None):
    """Preserve any existing cell borders; replace only sides you pass in."""
    cell.border = Border(
        left   = left   or cell.border.left,
        right  = right  or cell.border.right,
        top    = top    or cell.border.top,
        bottom = bottom or cell.border.bottom,
    )

def apply_subject_row_borders(
    ws,
    table_ref: str,
    *,
    subject_col_letter: str = "A",
    thin_color: str = "BFBFBF",    # light gray thin grid
    thin_style: str = "thin",
    thick_color: str = "000000",   # black for group separators
    thick_style: str = "medium",
):
    """
    Add a thin border around every data row within table_ref (e.g., 'A1:D200'),
    and add a thick border between subject groups (when subject in column A changes).

    - Assumes header is the first row of table_ref.
    - Subjects are contiguous because we wrote rows grouped by subject.
    """
    min_col, min_row, max_col, max_row = range_boundaries(table_ref)

    # Convert subject column letter to index (1-based)
    subject_col_idx = ord(subject_col_letter.upper()) - ord('A') + 1

    thin  = Side(style=thin_style,  color=thin_color)
    thick = Side(style=thick_style, color=thick_color)

    # 1) Apply a thin box around EVERY data row
    for r in range(min_row + 1, max_row + 1):  # skip header row
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            _merge_border(cell, top=thin, bottom=thin, left=thin, right=thin)

    # 2) Apply thick separator when Subject changes
    def subj_at(row_idx: int) -> str:
        return str(ws.cell(row=row_idx, column=subject_col_idx).value or "").strip()

    prev_subj = None
    for r in range(min_row + 1, max_row + 1):
        curr_subj = subj_at(r)

        # First data row: thick TOP to start first subject group
        if r == min_row + 1:
            for c in range(min_col, max_col + 1):
                _merge_border(ws.cell(r, c), top=thick)

        # Subject boundary: thick BOTTOM on previous row and thick TOP on current row
        if prev_subj is not None and curr_subj != prev_subj:
            for c in range(min_col, max_col + 1):
                _merge_border(ws.cell(r - 1, c), bottom=thick)  # end of previous subject group
                _merge_border(ws.cell(r,     c), top=thick)     # start of new subject group

        prev_subj = curr_subj

    # After loop: ensure the very last data row has a thick BOTTOM (end of last subject group)
    if max_row >= min_row + 1:
        for c in range(min_col, max_col + 1):
            _merge_border(ws.cell(max_row, c), bottom=thick)

def _merge_border(cell, *, top=None, bottom=None, left=None, right=None):
    """Preserve existing sides; only replace sides passed in."""
    cell.border = Border(
        left   = left   or cell.border.left,
        right  = right  or cell.border.right,
        top    = top    or cell.border.top,
        bottom = bottom or cell.border.bottom,
    )

def apply_table_borders(
    ws,
    table_ref: str,
    *,
    outer_color: str = "000000",    # black
    outer_style: str = "medium",    # thick-ish outer border
    inner_v_colors: list[str] | None = None,  # per-column right-edge colors
    inner_v_style: str = "thin",    # “little thin” separators
    inner_h_color: str | None = "D9D9D9",     # light gray horizontal (optional)
    inner_h_style: str = "hair",    # very light horizontals; use "thin" if you want stronger
):
    """
    Apply a solid outer border around the table and thin colored borders
    to separate each column.

    - ws: worksheet
    - table_ref: e.g., "A1:D57" (use the same ref you passed to the Table)
    - inner_v_colors: a list of hex colors (without '#'). One color per
      separator between columns. If shorter than required, colors cycle.
    """
    min_col, min_row, max_col, max_row = range_boundaries(table_ref)

    # Outer border sides
    outer = Side(style=outer_style, color=outer_color)
    top_side = outer
    bottom_side = outer
    left_side = outer
    right_side = outer

    # 1) Outer rectangle
    # Top edge
    for c in range(min_col, max_col + 1):
        _merge_border(ws.cell(min_row, c), top=top_side)
    # Bottom edge
    for c in range(min_col, max_col + 1):
        _merge_border(ws.cell(max_row, c), bottom=bottom_side)
    # Left edge
    for r in range(min_row, max_row + 1):
        _merge_border(ws.cell(r, min_col), left=left_side)
    # Right edge
    for r in range(min_row, max_row + 1):
        _merge_border(ws.cell(r, max_col), right=right_side)

    # 2) Internal vertical separators (color per column)
    # There are (max_col - min_col) internal vertical lines (right edge of each col, except last)
    if inner_v_colors is None:
        inner_v_colors = ["1F4E79", "548235", "C00000", "9F4C7C", "833C0C"]  # sample palette
    for c in range(min_col, max_col):  # up to second-last column
        color = inner_v_colors[(c - min_col) % len(inner_v_colors)]
        side = Side(style=inner_v_style, color=color)
        for r in range(min_row, max_row + 1):
            _merge_border(ws.cell(r, c), right=side)

    # 3) Optional: internal horizontal separators (light gray)
    if inner_h_color:
        h_side = Side(style=inner_h_style, color=inner_h_color)
        # Between header row and first data row:
        for c in range(min_col, max_col + 1):
            _merge_border(ws.cell(min_row, c), bottom=h_side)
        # Between data rows:
        for r in range(min_row + 1, max_row):  # interior rows only
            for c in range(min_col, max_col + 1):
                _merge_border(ws.cell(r, c), bottom=h_side)

def is_hindi_text(s: str) -> bool:
    """Heuristic: presence of Devanagari code points."""
    return any('\u0900' <= ch <= '\u097F' for ch in s)

def slugify_en(title: str) -> str:
    """Uppercase, keep alnum + spaces -> hyphens; collapse repeats; trim."""
    s = re.sub(r"[^A-Za-z0-9 ]+", " ", title)
    s = re.sub(r"\s+", "-", s.strip())
    s = re.sub(r"-{2,}", "-", s)
    return s.upper().strip("-")

def slugify_hi(title: str) -> str:
    """Keep Devanagari; whitespace/punct -> hyphen; collapse repeats; trim."""
    s = re.sub(r"[\s:–—_,.;!?()\[\]/+|]", "-", title)
    s = re.sub(r"-{2,}", "-", s)
    return s.strip("-")

def parse_chapter(raw: str, fallback_no: int) -> tuple[int, str]:
    """
    Accept 'LESSON-XX: Title' or 'Title'.
    Return (chapter_no, title). If no pattern, use fallback_no.
    """
    m = re.match(r"^\s*LESSON[- ]?(\d+)\s*:\s*(.+)$", raw, flags=re.IGNORECASE)
    if m:
        return int(m.group(1)), m.group(2).strip()
    return fallback_no, raw.strip()

def to_md_slug(num: int, title: str) -> str:
    slug = slugify_hi(title) if is_hindi_text(title) else slugify_en(title)
    return f"CHAPTER-{num:02d}-{slug}"

def join_chapter_slugs(chapters: List[str]) -> str:
    """Return newline-separated list of CHAPTER-XX-... slugs."""
    lines = []
    for idx, raw in enumerate(chapters, start=1):
        num, title = parse_chapter(raw, idx)
        lines.append(to_md_slug(num, title))
    return "\n".join(lines)

def autosize_columns(ws):
    """Approximate auto-fit; make Column B wider for multi-line slugs."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        width = min(120, max(12, max_len + 2))
        if letter == "B":  # Chapter Index
            width = max(width, 60)
        ws.column_dimensions[letter].width = width
        

def safe_save_xlsx(wb, target_path: Path, attempts: int = 6, delay: float = 0.5):
    """
    Save workbook via temp file & atomic replace with retries.
    Leaves a _LOCAL copy if final replace fails due to locks.
    """
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_name = tmp.name
    try:
        wb.save(tmp_name)
    except Exception:
        try:
            os.remove(tmp_name)
        except Exception:
            pass
        raise

    last_err = None
    for _ in range(attempts):
        try:
            if target_path.exists():
                try: os.chmod(target_path, 0o666)
                except Exception: pass
                try: target_path.unlink()
                except Exception: pass
            os.replace(tmp_name, target_path)  # atomic on Windows
            return
        except PermissionError as e:
            last_err = e
            time.sleep(delay)
            delay = min(2.5, delay * 1.7)
        except Exception as e:
            last_err = e
            break
    # Fallback if still locked
    fallback = target_path.parent / (target_path.stem + "_LOCAL.xlsx")
    try:
        os.replace(tmp_name, fallback)
    except Exception:
        raise RuntimeError(
            f"Could not replace '{target_path}'. Temp copy left at '{tmp_name}'."
        ) from last_err
    raise RuntimeError(
        f"Could not write to '{target_path}'. A good copy is saved at '{fallback}'."
    ) from last_err

def make_table_name(base: str, used: set[str]) -> str:
    """Unique table name; Excel requires alnum/underscore; cap length."""
    base = re.sub(r"[^A-Za-z0-9_]", "_", base)[:20] or "TBL"
    if base not in used:
        used.add(base)
        return base
    for i in range(1, 10000):
        cand = f"{base}_{i}"
        if cand not in used:
            used.add(cand)
            return cand
    raise RuntimeError("Unable to build unique table name.")

# ------------------------------ CORE ----------------------------------

def build_class_sheet(
    wb,
    class_num: int,
    class_data: Dict[str, List[str]],
    used_table_names: set[str],
    table_style: str,
):
    """
    Create one worksheet for a class with a single table:
    Subject | Chapter Index | Status | Remark

    This version writes ONE ROW PER CHAPTER.
    Column A repeats the subject for each chapter row.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation

    sheet_title = f"Class {class_num:02d}"
    ws = wb.create_sheet(title=sheet_title)

    # ---- Header ----
    headers = ["Subject", "Chapter Index", "Status", "Remark"]
    ws.append(headers)

    # ---- Rows: one row per chapter ----
    rows_added = 0
    if class_data:
        for subject in sorted(class_data.keys()):
            chapters = class_data[subject] or []
            if chapters:
                for idx, raw in enumerate(chapters, start=1):
                    num, title = parse_chapter(raw, idx)       # uses existing helper
                    md = to_md_slug(num, title)                 # CHAPTER-XX-... (existing helper)
                    ws.append([subject, md, "Not_start", ""])
                    rows_added += 1
            else:
                # Subject exists but no chapters -> single placeholder row
                ws.append([subject, "", "Not_start", ""])
                rows_added += 1
    else:
        # Class has no subjects -> single placeholder row so table is valid
        ws.append(["—", "", "Not_start", ""])
        rows_added += 1

    # ---- Header styling ----
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Wrap & top-align Chapter Index (Column A, B) for readability
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")  # Subject
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")  # Chapter Index


    # Ensure Status = Not_start and align to top-middle for every data row
    for r in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=r, column=3)  # Column C = Status
        if not status_cell.value:
            status_cell.value = "Not_start"
        status_cell.alignment = Alignment(horizontal="center", vertical="top")


    # Freeze header
    ws.freeze_panes = "A2"

    # ---- Status drop-down (Not_start, working, complete) ----
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_CHOICES)}"',
        allow_blank=True,
        showDropDown=True,
        errorTitle="Invalid choice",
        error="Select one of: Not_start, working, complete",
        promptTitle="Status",
        prompt="Pick a value",
    )
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        dv.add(f"C2:C{ws.max_row}")

    # ---- Excel Table over the used range ----
    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    tname = make_table_name(f"TBL_CLASS_{class_num:02d}", used_table_names)
    table = Table(displayName=tname, ref=ref)
    style = TableStyleInfo(
        name=table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,  # recommended off so borders/fills are visible
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Thin border on each row + thick separators when Subject changes
    apply_subject_row_borders(
        ws,
        ref,                           # e.g., "A1:D157"
        subject_col_letter="A",        # Subject is in column A
        thin_color="BFBFBF",           # thin grid color (adjust if you want)
        thin_style="thin",
        thick_color="000000",
        thick_style="medium",          # or "thick" for stronger group separators
    )


    # set table border code
    apply_table_borders(
        ws,
        ref,
        outer_color="000000",          # black outer
        outer_style="medium",          # solid outer frame
        inner_v_colors=["1F4E79", "548235", "C00000"],  # A|B|C separators; repeats if more cols
        inner_v_style="thin",          # “little thin”
        inner_h_color="D9D9D9",        # soft gray row separators (optional)
        inner_h_style="hair",
    )

    # ---- Column widths (narrower Chapter Index, still wrapped) ----
    ws.column_dimensions['A'].width = 12  # Subject
    ws.column_dimensions['B'].width = 52  # Chapter Index (smaller; each row has one slug)
    ws.column_dimensions['C'].width = 14  # Status
    ws.column_dimensions['D'].width = 54  # Remark



    from openpyxl.styles import PatternFill
    def subject_color(subject: str) -> str:
        return SUBJECT_FILL.get(subject, SUBJECT_FILL["_default"])

    # Apply fill to the entire row A:D, based on Subject in column A
    for r in range(2, ws.max_row + 1):
        subj = (ws.cell(row=r, column=1).value or "").strip()
        color_hex = subject_color(subj)
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        for c in range(1, 5):  # columns A..D
            ws.cell(row=r, column=c).fill = fill



def build_workbook(out_path: Path, table_style: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_table_names: set[str] = set()

    # Ensure sheets for Class 01 .. Class 12 (even if some classes have empty dicts)
    for class_num in range(1, 13):
        class_key = f"Class {class_num}"
        class_data = CURRICULUM.get(class_key, {})
        build_class_sheet(wb, class_num, class_data, used_table_names, table_style)

    safe_save_xlsx(wb, out_path)

# ------------------------------ CLI -----------------------------------

def parse_args():
    p = argparse.ArgumentParser(description="Build NCERT workbook: one sheet per class; all subjects per sheet.")
    p.add_argument("--out", type=str, default=str(DEFAULT_OUT_DIR), help="Output directory (default: script folder)")
    p.add_argument("--file", type=str, default=DEFAULT_FILENAME, help="Output filename (default: NCERT_Classes.xlsx)")
    p.add_argument("--style", type=str, default=DEFAULT_STYLE, help=f"Excel table style (default: {DEFAULT_STYLE})")
    p.add_argument("--unique", action="store_true", help="Append timestamp to filename to avoid locked replaces")
    return p.parse_args()

def main():
    args = parse_args()
    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / args.file
    if args.unique:
        stem, suf = out_path.stem, out_path.suffix or ".xlsx"
        out_path = out_dir / f"{stem}_{datetime.now():%Y%m%d_%H%M%S}{suf}"

    build_workbook(out_path, args.style)
    print(f"[OK] Saved workbook: {out_path}")

if __name__ == "__main__":
    main()
